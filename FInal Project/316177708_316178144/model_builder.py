"""
    Barak Bonker - 316177708
    Amit Avigdor -  316178144
"""

import os
import tkinter as tk
import numpy as np
from windowSettings import *
from tkinter import filedialog
import pandas as pd
import model_components as mc
import pickle
import openpyxl
from sklearn import metrics


class MainScreen(tk.Tk):
    """
        the class for the model builder and runner frontend and partly backend.
        shows a gui for the user to choose the csv file how to clean it and the model to run.
        the software creates clean csv, model file, runs the model
        saves the results as png file of confusion matrix and finally on Excel map.
    """
    def __init__(self, *args, **kwargs):
        """
            init and fill the window with the components
        """
        super().__init__(*args, **kwargs)

        # set the window
        set_window(self)
        self.protocol("WM_DELETE_WINDOW", self.destroy)

        # usefull dictionaries for later use
        self.fill_dict = {"All data": mc.fillByAll, "Classify value": mc.fillByClass}
        self.model_dict = {"Decision tree": mc.skDecisionTree, "Naive Bayes": mc.skNaiveBayes, "KNN": mc.KNN,
                      "K-means": mc.kMeans}

        # choosing the csv file label and button
        ttk.Label(self, text="Choose .csv file: ").grid(row=0, column=0, padx=10, pady=20)
        self.browse_button = ttk.Button(self, text="Browse Files", command=self.browseFiles)
        self.browse_button.grid(row=0, column=1, padx=20, ipadx=20, ipady=5)

        # choosing the classify column label and combobox
        self.column_selected = tk.StringVar()
        ttk.Label(self, text="Choose classify column: ").grid(row=1, column=0, padx=10, pady=20)
        self.classify_combox_selection = ttk.Combobox(self, textvariable=self.column_selected, state="disable")
        self.classify_combox_selection.grid(row=1, column=1, padx=20)

        # choosing how to fill the missing values label and combobox
        self.missing_value_selected = tk.StringVar()
        ttk.Label(self, text="Choose missing values filling method: ").grid(row=2, column=0, padx=10, pady=20)
        self.missing_values_combox_selection = ttk.Combobox(self, textvariable=self.missing_value_selected,
                                                            state="readonly")
        self.missing_values_combox_selection.grid(row=2, column=1, padx=20)
        self.missing_values_combox_selection["values"] = ["All data", "Classify value"]

        # choosing if to normalize or not label and combobox
        self.normalization_selected = tk.StringVar()
        ttk.Label(self, text="Choose if normalization needed: ").grid(row=3, column=0, padx=10, pady=20)
        self.normalization_combox_selection = ttk.Combobox(self, textvariable=self.normalization_selected,
                                                           state="readonly")
        self.normalization_combox_selection.grid(row=3, column=1, padx=20)
        self.normalization_combox_selection["values"] = ["Yes", "No"]

        # choosing the discretisize methode label and combobox
        self.discretization_selected = tk.StringVar()
        ttk.Label(self, text="Choose if discretization needed: ").grid(row=4, column=0, padx=10, pady=20)
        self.discretization_combox_selection = ttk.Combobox(self, textvariable=self.discretization_selected,
                                                            state="readonly")
        self.discretization_combox_selection.grid(row=4, column=1, padx=20)
        self.discretization_combox_selection["values"] = ["Without", "Equal-frequency", "Equal-width", "Entropy-based"]

        # choosing how many bins methode label and entry
        self.bins_selected = tk.StringVar()
        ttk.Label(self, text="How many bins (if needed): ").grid(row=5, column=0, padx=10, pady=20)
        self.bins_entry_selection = ttk.Entry(self, width=20, textvariable=self.bins_selected)
        self.bins_entry_selection.grid(row=5, column=1, padx=20)

        # choosing the model label and combobox
        self.model_selected = tk.StringVar()
        ttk.Label(self, text="Choose algorithm model: ").grid(row=6, column=0, padx=10, pady=20)
        self.model_combox_selection = ttk.Combobox(self, textvariable=self.model_selected, state="readonly")
        self.model_combox_selection.grid(row=6, column=1, padx=20)
        self.model_combox_selection["values"] = ["Decision tree", "Naive Bayes", "KNN", "K-means"]
        self.model_combox_selection.bind("<<ComboboxSelected>>", self.onModelSelect)

        # choosing the model implement methode label and combobox
        self.implement_selected = tk.StringVar()
        ttk.Label(self, text="Choose model implementation: ").grid(row=7, column=0, padx=10, pady=20)
        self.model_combox_implement = ttk.Combobox(self, textvariable=self.implement_selected, state="disabled")
        self.model_combox_implement.grid(row=7, column=1, padx=20)
        self.model_combox_implement["values"] = ["Self made", "Sklearn"]

        # button to clean the csv file and build the model (save them both)
        self.build_button = ttk.Button(self, text="Clean File & Build Model", command=self.cleanAndBuildModel,
                                       state="disable")
        self.build_button.grid(row=8, column=0, ipadx=20, ipady=5, pady=20)

        # button to run the model and save the results
        self.run_button = ttk.Button(self, text="Run Model", command=self.runModel, state="disable")
        self.run_button.grid(row=8, column=1, ipadx=20, ipady=5, pady=20)


    def browseFiles(self):
        """
            open a file dialog to choose the csv file.
            open the chosen file and if not empty, get the columns names for later use.
        """
        def emptyFileError():
            """
                error message if the file is empty.
            """
            def OK_to_error():
                empty_file_error_window.destroy()
                self.browse_button["state"] = "normal"
                self.build_button["state"] = "disable"
            # setting up the error window
            empty_file_error_window = tk.Tk()
            set_window(empty_file_error_window)
            empty_file_error_window.title("Error")
            self.browse_button["state"] = "disable"
            ttk.Label(empty_file_error_window, text="The file you imported is empty, cant work with it. ").grid(row=0,
                                                                                                                column=0,
                                                                                                                padx=30,
                                                                                                                pady=10)
            ttk.Label(empty_file_error_window, text="please choose another file. ").grid(row=2, column=0, padx=30)
            ttk.Button(empty_file_error_window, text="OK", command=OK_to_error).grid(row=3, column=0, ipadx=5,
                                                                                     ipady=2, padx=5, pady=20)
            empty_file_error_window.protocol("WM_DELETE_WINDOW", OK_to_error)

        # open the file dialog and get the path of the chosen file
        self.filePath = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                                   filetypes=(("csv files", "*.csv*"),))
        if self.filePath is not "":
            # if the path is not empty, try to open the file and get the columns names
            try:
                self.file = pd.read_csv(self.filePath)
                self.classify_combox_selection["state"] = "readonly"
                self.classify_combox_selection["values"] = self.file.columns.to_list()
                self.build_button["state"] = "normal"
                self.run_button["state"] = "disable"
            except pd.errors.EmptyDataError:
                # if the file is empty, show the error message
                emptyFileError()
                self.classify_combox_selection.set('')
                self.classify_combox_selection["state"] = "disable"
                self.build_button["state"] = "disable"
                self.run_button["state"] = "disable"
        else:
            # if the path is empty, disable buttons
            self.classify_combox_selection.set('')
            self.classify_combox_selection["state"] = "disable"
            self.build_button["state"] = "disable"
            self.run_button["state"] = "disable"


    def onModelSelect(self, x=None):
        """
            when the model is selected at the combobox, if needed enable the implementation combox.
        """
        if self.model_combox_selection.get() == "Naive Bayes":
            self.model_combox_implement["state"] = "readonly"
        else:
            self.model_combox_implement["state"] = "disable"
            self.implement_selected.set('Sklearn')


    def cleanAndBuildModel(self):
        """
            clean the csv file and build the model.
            save the model and the clean csv file as new files.
        """
        if self.column_selected.get() == "" == self.missing_value_selected.get() == "" or self.normalization_selected.get() == "" \
                or not self.isDiscretAndBinsOk() or self.model_selected.get() == "" or self.implement_selected.get() == "":
            # if one of the required fields is empty, show the error message
            def OK_to_error():
                fill_all_window.destroy()
                self.build_button["state"] = "normal"
            # setting up the error window
            fill_all_window = tk.Tk()
            set_window(fill_all_window)
            fill_all_window.title("Error")
            self.build_button["state"] = "disable"
            ttk.Label(fill_all_window, text="Please fill all the fields. ").grid(row=0, column=0, padx=30, pady=10)
            ttk.Button(fill_all_window, text="OK", command=OK_to_error).grid(row=3, column=0, ipadx=5,
                                                                             ipady=2, padx=5, pady=20)
            fill_all_window.protocol("WM_DELETE_WINDOW", OK_to_error)
        else:
            # if all the fields are filled, clean the csv file and build the model and unable the run model button
            self.cleanFile()
            self.buildModel()


    def isDiscretAndBinsOk(self):
        """
            check if discretization is chosen and if bins is inserted (if not chosen without discretization).
        :return: if the discretization and bins selected is valid or not.
        """
        if self.discretization_selected.get() == "Without":
            return True
        elif self.discretization_selected.get() == "Equal-frequency" or self.discretization_selected.get() == "Equal-width"\
                or self.discretization_selected.get() == "Entropy-based":
            return self.bins_selected.get().isnumeric()
        else:
            return False


    def cleanFile(self):
        """
            clean the csv file by the choices of the user and save the clean csv it as a new file.
        """
        mc.clean(self.file, self.column_selected.get())
        self.fill_dict[self.missing_value_selected.get()](self.file, self.column_selected.get())
        if self.normalization_selected.get() == "Yes":
            mc.normalize(self.file, self.column_selected.get())
        if self.discretization_selected.get() != "Without":
            mc.discretize(self.file, self.column_selected.get(), self.discretization_selected.get(),
                          self.bins_selected.get())
        # save the clean csv file as a new csv file with the same name but with added _clean to the name
        self.file.to_csv(os.path.basename(self.filePath).split(".")[0] + "_clean.csv", index=False)
        self.class_values = self.file[self.column_selected.get()].unique()


    def buildModel(self):
        """
            encode the data, split it to train and test.
            build the model by the choices of the user and save it as a new file using pickle.
        """
        try:
            # try to encode the data, split and build the model
            data_cols, classify_col = mc.encodeAndPopClass(self.file, self.column_selected.get())
            self.data_train, self.data_test, self.class_train, self.class_test = mc.SplitTrainTest(data_cols, classify_col)

            # runs the model that the user chose
            if self.model_selected.get() == "Decision tree":
                model = mc.skDecisionTree(self.data_train, self.class_train)
            elif self.model_selected.get() == "Naive Bayes":
                if self.implement_selected.get() == "Sklearn":
                    model = mc.skNaiveBayes(self.data_train, self.class_train)
                else:
                    model = mc.selfNaiveBayes(self.data_train, self.class_train)
            elif self.model_selected.get() == "KNN":
                model = mc.KNN(self.data_train, self.class_train)
            else:
                model = mc.kMeans(self.data_train, self.class_train)

            # save the model as a new binary file with the name "model.pkl"
            with open('model.pkl', 'wb') as f:
                pickle.dump(model, f)

            self.run_button["state"] = "normal"

        except (ValueError, TypeError):
            # if the model is not built, show the error message
            def OK_to_error():
                fill_all_window.destroy()
                self.build_button["state"] = "normal"

            # setting up the error window
            fill_all_window = tk.Tk()
            set_window(fill_all_window)
            fill_all_window.title("Error")
            self.build_button["state"] = "disable"
            ttk.Label(fill_all_window, text="model could not be built, try again with other choices. ").grid(row=0, column=0, padx=30, pady=10)
            ttk.Button(fill_all_window, text="OK", command=OK_to_error).grid(row=3, column=0, ipadx=5,
                                                                             ipady=2, padx=5, pady=20)
            fill_all_window.protocol("WM_DELETE_WINDOW", OK_to_error)

    def runModel(self):
        """
            open and run the model and save the results as pngs of the confusion Matrix and creates the majority law array
        """
        model = pickle.load(open('model.pkl', 'rb'))
        self.train_prediction, self.test_prediction = mc.predict(model, self.data_train, self.data_test, self.implement_selected.get())
        mc.confusionMatrix(self.train_prediction, self.test_prediction, self.class_train, self.class_test, self.model_selected.get(), self.class_values)
        self.majority_law_array = np.full((1, len(self.class_test)), self.class_test.mode(), dtype=int)
        self.fillResults()


    def fillResults(self):
        """
            fill the results Excel file with the results of the model and the cleaning settings.
        """
        try:
            #  try open the results Excel file
            wb = openpyxl.load_workbook("results.xlsx")
        except FileNotFoundError:
            # if the file doesn't exist, create a new one and fill the columns
            wb = openpyxl.Workbook()
            ws = wb.worksheets[0]
            ws.append(
                ["file name", "classify name", "filling methode", "normalization", "discretization", "bins", "model",
                 "confusion matrix train", "confusion matrix test", "majority law accuracy"])
            ws.column_dimensions['H'].width = 42
            ws.column_dimensions['I'].width = 42
            wb.save("results.xlsx")

        # open the results Excel file
        wb = openpyxl.load_workbook("results.xlsx")
        ws = wb.worksheets[0]

        # fill the results Excel file with the results of the model and the cleaning settings to a new row
        ws.append([os.path.basename(self.filePath), self.column_selected.get(), self.missing_value_selected.get(),
                   self.normalization_selected.get(), self.discretization_selected.get(), self.bins_selected.get(),
                   self.model_selected.get(), "", "", metrics.accuracy_score(self.class_test, self.majority_law_array[0])])

        # adds the confusion matrix to the Excel file and saves the Excel file
        ws.row_dimensions[ws.max_row].height = 220
        img = openpyxl.drawing.image.Image('Confusion_Matrix_train.png')
        ws.add_image(img, 'H'+str(ws.max_row))
        img = openpyxl.drawing.image.Image('Confusion_Matrix_test.png')
        ws.add_image(img, 'I' + str(ws.max_row))
        wb.save("results.xlsx")


MainScreen().mainloop()


